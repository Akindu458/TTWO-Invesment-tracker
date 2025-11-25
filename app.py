import streamlit as st
import yfinance as yf
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import pandas as pd

st.title("TTWO Investment Tracker (USD Only)")

# --- Investor Input ---
name = st.text_input("Enter investor name")
amount_usd = st.number_input("Amount invested (USD)", min_value=0.0, step=1.0)

# --- Functions ---
def get_ttwo_price():
    """Fetches the latest TTWO stock price safely"""
    try:
        stock = yf.Ticker("TTWO")
        price = stock.history(period="1d")["Close"].iloc[-1]
        return float(price)
    except Exception as e:
        st.error(f"Error fetching TTWO stock price: {e}")
        return None

def save_to_excel(row):
    """Saves investment data to Excel"""
    filename = "TTWO Investment.xlsx"
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Investments"
        headers = [
            "Timestamp",
            "Name",
            "Amount Invested (USD)",
            "TTWO Stock Price (USD)",
            "Number of Shares"
        ]
        ws.append(headers)
        wb.save(filename)

    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row)
    wb.save(filename)
    
    # Open Excel automatically
    try:
        os.startfile(os.path.abspath(filename))
    except Exception as e:
        st.warning(f"Could not open Excel automatically: {e}")

def load_investments():
    """Load investment data from Excel"""
    filename = "TTWO Investment.xlsx"
    if os.path.exists(filename):
        df = pd.read_excel(filename)
        return df
    else:
        return pd.DataFrame()

# --- Process Investment ---
if st.button("Submit Investment"):
    if not name or amount_usd <= 0:
        st.warning("Please enter a valid name and investment amount.")
    else:
        ttwo_price = get_ttwo_price()
        if ttwo_price is None:
            st.stop()

        shares = amount_usd / ttwo_price

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = [
            timestamp,
            name,
            round(amount_usd, 2),
            round(ttwo_price, 4),
            round(shares, 6)
        ]
        save_to_excel(row)
        st.success(f"Saved! {shares:.4f} TTWO shares can be purchased at ${ttwo_price:.2f}")

# --- Display past investments ---
st.subheader("All Investments")
df_investments = load_investments()
if not df_investments.empty:
    df_investments.index = range(1, len(df_investments)+1)
    st.dataframe(df_investments)

    total_invested = df_investments["Amount Invested (USD)"].sum()
    total_shares = df_investments["Number of Shares"].sum()
    st.subheader("Summary")
    st.markdown(f"- **Total Invested:** ${total_invested:,.2f}")
    st.markdown(f"- **Total TTWO Shares:** {total_shares:,.6f}")
else:
    st.info("No investments yet.")
